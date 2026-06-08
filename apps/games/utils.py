import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.utils import timezone
from django.http import HttpResponse
from collections import defaultdict


ERROR_TYPE_TIME_FORMAT = 'time_format'
ERROR_TYPE_PLAYER_MISSING = 'player_missing'
ERROR_TYPE_SCORE_IMBALANCE = 'score_imbalance'
ERROR_TYPE_INSUFFICIENT_PLAYERS = 'insufficient_players'
ERROR_TYPE_SCORE_FORMAT = 'score_format'
ERROR_TYPE_PARSE_ERROR = 'parse_error'

ERROR_TYPE_LABELS = {
    ERROR_TYPE_TIME_FORMAT: '时间格式异常',
    ERROR_TYPE_PLAYER_MISSING: '玩家缺失',
    ERROR_TYPE_SCORE_IMBALANCE: '分数不平衡',
    ERROR_TYPE_INSUFFICIENT_PLAYERS: '玩家数量不足',
    ERROR_TYPE_SCORE_FORMAT: '得分格式错误',
    ERROR_TYPE_PARSE_ERROR: '解析错误',
}


class ImportErrorItem:
    def __init__(self, row_num, message, error_type, details=None):
        self.row_num = row_num
        self.message = message
        self.error_type = error_type
        self.details = details or {}

    def to_dict(self):
        return {
            'row_num': self.row_num,
            'message': self.message,
            'error_type': self.error_type,
            'error_type_label': ERROR_TYPE_LABELS.get(self.error_type, '其他错误'),
            'details': self.details,
        }


class ImportResult:
    def __init__(self):
        self.games_data = []
        self.errors = []
        self._error_groups = defaultdict(list)

    def add_error(self, error_item):
        self.errors.append(error_item)
        self._error_groups[error_item.error_type].append(error_item)

    @property
    def success_count(self):
        return len(self.games_data)

    @property
    def error_count(self):
        return len(self.errors)

    @property
    def total_count(self):
        return self.success_count + self.error_count

    def get_errors_by_type(self, error_type):
        return self._error_groups.get(error_type, [])

    def get_error_summary(self):
        groups = []
        for error_type, label in ERROR_TYPE_LABELS.items():
            errors = self._error_groups.get(error_type, [])
            if errors:
                groups.append({
                    'error_type': error_type,
                    'label': label,
                    'count': len(errors),
                    'errors': [e.to_dict() for e in errors],
                })
        other_errors = [
            e for e in self.errors
            if e.error_type not in ERROR_TYPE_LABELS
        ]
        if other_errors:
            groups.append({
                'error_type': 'other',
                'label': '其他错误',
                'count': len(other_errors),
                'errors': [e.to_dict() for e in other_errors],
            })
        return groups

    def get_summary_dict(self):
        return {
            'total': self.total_count,
            'success_count': self.success_count,
            'error_count': self.error_count,
            'error_groups': self.get_error_summary(),
        }

    def get_flat_errors(self):
        return [e.to_dict() for e in self.errors]

    def get_error_messages(self):
        return [e.message for e in self.errors]


def export_games_to_pdf(games_qs, view_mode='score'):
    """Export games queryset to PDF file using reportlab"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    # Try to register a CJK font; fall back to built-in if unavailable
    font_name = 'Helvetica'
    cjk_fonts = [
        # Debian/Ubuntu (installed via fonts-wqy-microhei)
        '/usr/share/fonts/truetype/wqy/wqy-microhei.ttc',
        '/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc',
        # Debian Noto CJK
        '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
        '/usr/share/fonts/noto-cjk/NotoSansCJKsc-Regular.otf',
        # macOS
        '/System/Library/Fonts/PingFang.ttc',
        '/System/Library/Fonts/STHeiti Light.ttc',
        # Windows
        'C:/Windows/Fonts/msyh.ttc',
        'C:/Windows/Fonts/simhei.ttf',
        'C:/Windows/Fonts/simsun.ttc',
    ]
    for fp in cjk_fonts:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont('CJK', fp))
                font_name = 'CJK'
                break
            except Exception:
                pass

    if font_name == 'Helvetica':
        # No CJK font available – return a plain text notice
        response = HttpResponse(
            '系统未安装中文字体，无法导出 PDF。请联系管理员在服务器上安装 fonts-wqy-microhei 后重试。',
            content_type='text/plain; charset=utf-8'
        )
        return response

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', fontName=font_name, fontSize=16, spaceAfter=8,
                                  textColor=colors.HexColor('#c9a227'), alignment=1)
    subtitle_style = ParagraphStyle('Sub', fontName=font_name, fontSize=9, spaceAfter=12,
                                     textColor=colors.grey, alignment=1)
    cell_style = ParagraphStyle('Cell', fontName=font_name, fontSize=8)

    elements = []
    if view_mode == 'amount':
        title_text = '麻将战绩报告（金额）'
    else:
        title_text = '麻将战绩报告'
    elements.append(Paragraph(title_text, title_style))
    elements.append(Paragraph(
        f'导出时间：{timezone.now().strftime("%Y-%m-%d %H:%M")}  共 {games_qs.count()} 场',
        subtitle_style
    ))
    elements.append(Spacer(1, 4 * mm))

    if view_mode == 'amount':
        headers = ['#', '游戏时间', '地点', '类型', '底分(元/分)', '参与玩家 & 金额(元)', '备注']
    else:
        headers = ['#', '游戏时间', '地点', '类型', '底分', '参与玩家 & 得分', '备注']
    table_data = [headers]

    for i, game in enumerate(games_qs[:500], 1):
        players = game.players.select_related('user').order_by('-score')
        if view_mode == 'amount':
            player_str = '  '.join(
                f"{p.user.get_display_name()}:{'+' if p.score >= 0 else ''}{p.score * game.base_score}元"
                for p in players
            )
        else:
            player_str = '  '.join(
                f"{p.user.get_display_name()}:{'+' if p.score >= 0 else ''}{p.score}"
                for p in players
            )
        row = [
            str(i),
            game.game_time.strftime('%Y-%m-%d %H:%M'),
            game.location or '-',
            game.get_game_type_display(),
            str(game.base_score),
            Paragraph(player_str, cell_style),
            Paragraph(game.notes[:40] if game.notes else '-', cell_style),
        ]
        table_data.append(row)

    col_widths = [10*mm, 38*mm, 30*mm, 28*mm, 16*mm, 100*mm, 45*mm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3240')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#c9a227')),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f8fb')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
        ('ROWHEIGHT', (0, 0), (-1, 0), 14),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    if view_mode == 'amount':
        filename = f'mahjong_records_amount_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    else:
        filename = f'mahjong_records_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_games_to_excel(games_qs, view_mode='score'):
    """Export games queryset to Excel file"""
    wb = openpyxl.Workbook()
    ws = wb.active
    if view_mode == 'amount':
        ws.title = '麻将战绩(金额)'
    else:
        ws.title = '麻将战绩'

    # Styles
    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='1a3240', end_color='1a3240', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    if view_mode == 'amount':
        headers = ['游戏ID', '游戏时间', '地点', '游戏类型', '底分(元/分)', '参与玩家', '各玩家金额(元)', '是否补录', '备注']
    else:
        headers = ['游戏ID', '游戏时间', '地点', '游戏类型', '底分', '参与玩家', '各玩家得分', '是否补录', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 25

    for row, game in enumerate(games_qs, 2):
        players = game.players.select_related('user').order_by('-score')
        player_names = ', '.join([p.user.get_display_name() for p in players])
        if view_mode == 'amount':
            player_scores = ', '.join([f'{p.user.get_display_name()}:{p.score * game.base_score}' for p in players])
        else:
            player_scores = ', '.join([f'{p.user.get_display_name()}:{p.score}' for p in players])

        row_data = [
            game.pk,
            game.game_time.strftime('%Y-%m-%d %H:%M'),
            game.location or '',
            game.get_game_type_display(),
            game.base_score,
            player_names,
            player_scores,
            '是' if game.is_supplemental else '否',
            game.notes or '',
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            if row % 2 == 0:
                cell.fill = PatternFill(start_color='f0f8ff', end_color='f0f8ff', fill_type='solid')

    # Auto-width columns
    col_widths = [10, 18, 15, 15, 8, 25, 35, 10, 20]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    if view_mode == 'amount':
        filename = f'mahjong_records_amount_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    else:
        filename = f'mahjong_records_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def get_import_template():
    """Generate Excel import template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '战绩导入模板'

    headers = ['游戏时间(YYYY-MM-DD HH:MM)', '地点', '游戏类型', '底分', '玩家1用户名', '玩家1得分',
               '玩家2用户名', '玩家2得分', '玩家3用户名', '玩家3得分', '玩家4用户名', '玩家4得分', '备注']

    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='c9a227', end_color='c9a227', fill_type='solid')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    example_rows = [
        ['2024-01-15 19:30', '家里', 'mahjong_16', '1',
         'zhangsan', '30', 'lisi', '-10', 'wangwu', '-10', 'zhaoliu', '-10', '示例数据'],
        ['2024-01-16 20:00', '茶馆', 'mahjong_16', '2',
         'zhangsan', '-20', 'lisi', '40', 'wangwu', '-10', '', '', ''],
    ]
    for row_idx, example_row in enumerate(example_rows, 2):
        for col, value in enumerate(example_row, 1):
            ws.cell(row=row_idx, column=col, value=value)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="mahjong_import_template.xlsx"'
    wb.save(response)
    return response


def _parse_rows(rows_iter, start_row=2):
    """Parse iterable of row tuples into game data (shared by Excel and CSV parsers)

    Returns:
        ImportResult: 结构化的导入结果，包含成功数据和分类错误
    """
    result = ImportResult()
    from apps.accounts.models import User
    from datetime import datetime

    for row_num, row in enumerate(rows_iter, start_row):
        if not any(row):
            continue
        try:
            game_time_str = str(row[0]).strip() if row[0] else ''
            location = str(row[1]).strip() if row[1] else ''
            game_type = str(row[2]).strip() if row[2] else 'mahjong_16'
            base_score = int(row[3]) if row[3] else 1
            notes = str(row[12]).strip() if len(row) > 12 and row[12] else ''

            try:
                if isinstance(row[0], datetime):
                    game_time = timezone.make_aware(row[0])
                else:
                    game_time = timezone.make_aware(datetime.strptime(game_time_str, '%Y-%m-%d %H:%M'))
            except ValueError:
                result.add_error(ImportErrorItem(
                    row_num=row_num,
                    message=f'第{row_num}行：时间格式错误 "{game_time_str}"',
                    error_type=ERROR_TYPE_TIME_FORMAT,
                    details={'game_time_str': game_time_str},
                ))
                continue

            players_data = []
            total_score = 0
            has_score_format_error = False
            missing_players = []
            for i in range(4):
                username_col = 4 + i * 2
                score_col = 5 + i * 2
                if len(row) > score_col and row[username_col]:
                    username = str(row[username_col]).strip()
                    try:
                        score = int(row[score_col]) if row[score_col] is not None else 0
                    except (ValueError, TypeError):
                        result.add_error(ImportErrorItem(
                            row_num=row_num,
                            message=f'第{row_num}行：玩家{i+1}得分格式错误',
                            error_type=ERROR_TYPE_SCORE_FORMAT,
                            details={'player_index': i + 1, 'username': username},
                        ))
                        has_score_format_error = True
                        score = 0
                    try:
                        user = User.objects.get(username=username)
                        players_data.append({'user': user, 'score': score})
                        total_score += score
                    except User.DoesNotExist:
                        missing_players.append(username)
                        result.add_error(ImportErrorItem(
                            row_num=row_num,
                            message=f'第{row_num}行：用户 "{username}" 不存在',
                            error_type=ERROR_TYPE_PLAYER_MISSING,
                            details={'username': username},
                        ))

            if len(players_data) < 2:
                if not missing_players and not has_score_format_error:
                    result.add_error(ImportErrorItem(
                        row_num=row_num,
                        message=f'第{row_num}行：至少需要2名玩家',
                        error_type=ERROR_TYPE_INSUFFICIENT_PLAYERS,
                        details={'player_count': len(players_data)},
                    ))
                continue

            if total_score != 0:
                result.add_error(ImportErrorItem(
                    row_num=row_num,
                    message=f'第{row_num}行：得分总和不为0（当前为{total_score}）',
                    error_type=ERROR_TYPE_SCORE_IMBALANCE,
                    details={'total_score': total_score},
                ))
                continue

            result.games_data.append({
                'game_time': game_time,
                'location': location,
                'game_type': game_type,
                'base_score': base_score,
                'notes': notes,
                'players': players_data,
                'row_num': row_num,
            })
        except Exception as e:
            result.add_error(ImportErrorItem(
                row_num=row_num,
                message=f'第{row_num}行：解析错误 {str(e)}',
                error_type=ERROR_TYPE_PARSE_ERROR,
                details={'exception': str(e)},
            ))

    return result


def parse_import_file(file_obj):
    """Parse Excel or CSV import file, return ImportResult with structured data

    Returns:
        ImportResult: 结构化的导入结果，包含成功数据和分类错误
    """
    filename = getattr(file_obj, 'name', '')
    if filename.lower().endswith('.csv'):
        import csv
        try:
            content = file_obj.read()
            for enc in ('utf-8-sig', 'utf-8', 'gb18030', 'gbk'):
                try:
                    text = content.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                result = ImportResult()
                result.add_error(ImportErrorItem(
                    row_num=0,
                    message='CSV文件编码无法识别，请另存为UTF-8编码后再导入',
                    error_type=ERROR_TYPE_PARSE_ERROR,
                ))
                return result
            reader = csv.reader(text.splitlines())
            rows = list(reader)
            if not rows:
                result = ImportResult()
                result.add_error(ImportErrorItem(
                    row_num=0,
                    message='CSV文件为空',
                    error_type=ERROR_TYPE_PARSE_ERROR,
                ))
                return result
            data_rows = [tuple(r) for r in rows[1:]]
            return _parse_rows(iter(data_rows))
        except Exception as e:
            result = ImportResult()
            result.add_error(ImportErrorItem(
                row_num=0,
                message=f'CSV文件解析失败：{str(e)}',
                error_type=ERROR_TYPE_PARSE_ERROR,
            ))
            return result

    try:
        wb = openpyxl.load_workbook(file_obj)
        ws = wb.active
    except Exception as e:
        result = ImportResult()
        result.add_error(ImportErrorItem(
            row_num=0,
            message=f'文件格式错误：{str(e)}',
            error_type=ERROR_TYPE_PARSE_ERROR,
        ))
        return result

    return _parse_rows(ws.iter_rows(min_row=2, values_only=True))


def apply_game_filters(queryset, form):
    """Apply filter form to games queryset"""
    if not form.is_valid():
        return queryset

    from django.utils import timezone
    from datetime import timedelta, datetime

    data = form.cleaned_data
    date_range = data.get('date_range')
    now = timezone.now()

    if date_range == 'today':
        queryset = queryset.filter(game_time__date=now.date())
    elif date_range == 'week':
        week_start = now - timedelta(days=now.weekday())
        queryset = queryset.filter(game_time__gte=week_start.replace(hour=0, minute=0, second=0))
    elif date_range == 'month':
        queryset = queryset.filter(game_time__year=now.year, game_time__month=now.month)
    elif date_range == 'custom':
        if data.get('date_from'):
            queryset = queryset.filter(game_time__date__gte=data['date_from'])
        if data.get('date_to'):
            queryset = queryset.filter(game_time__date__lte=data['date_to'])

    if data.get('player'):
        queryset = queryset.filter(players__user=data['player'])

    if data.get('game_type'):
        queryset = queryset.filter(game_type=data['game_type'])

    if data.get('is_supplemental') != '':
        is_supp = data.get('is_supplemental') == '1'
        if data.get('is_supplemental'):
            queryset = queryset.filter(is_supplemental=is_supp)

    if data.get('search'):
        from django.db.models import Q
        search = data['search']
        queryset = queryset.filter(
            Q(location__icontains=search) | Q(notes__icontains=search)
        )

    return queryset


def get_playmate_stats(player_id, date_from=None, date_to=None, room_id=None,
                       sort_by='game_count', order='desc', limit=20):
    """计算玩家的同桌关系统计

    Args:
        player_id: 目标玩家ID
        date_from: 开始日期 (datetime)，可选
        date_to: 结束日期 (datetime)，可选
        room_id: 房间ID，可选，限定统计范围
        sort_by: 排序字段 (game_count/avg_score/win_rate/total_score/last_played)，默认 game_count
        order: 排序方向 (asc/desc)，默认 desc
        limit: 返回数量限制，默认 20

    Returns:
        list: 同桌玩家统计列表，每项包含：
            - teammate_id: 同桌玩家ID
            - teammate_name: 同桌玩家显示名称
            - game_count: 共同参战次数
            - avg_score: 同桌时目标玩家的平均得分
            - win_rate: 同桌时目标玩家的胜率（百分比）
            - wins: 同桌时获胜次数
            - losses: 同桌时失败次数
            - total_score: 同桌时总得分
            - last_played: 最后一次同桌时间
    """
    from .models import GamePlayer
    from django.db.models import Count, Avg, Sum, Max, Q
    from apps.accounts.models import User

    try:
        player = User.objects.get(pk=player_id)
    except User.DoesNotExist:
        return []

    my_games_qs = GamePlayer.objects.filter(
        user=player,
        game__status='completed'
    ).select_related('game')

    if date_from:
        my_games_qs = my_games_qs.filter(game__game_time__gte=date_from)
    if date_to:
        my_games_qs = my_games_qs.filter(game__game_time__lte=date_to)
    if room_id:
        my_games_qs = my_games_qs.filter(game__room_id=room_id)

    my_game_ids = list(my_games_qs.values_list('game_id', flat=True))

    if not my_game_ids:
        return []

    teammate_qs = GamePlayer.objects.filter(
        game_id__in=my_game_ids
    ).exclude(
        user_id=player_id
    ).select_related('user', 'game')

    my_scores_by_game = {gp.game_id: gp for gp in my_games_qs}

    stats_by_teammate = {}
    for gp in teammate_qs:
        uid = gp.user_id
        if uid not in stats_by_teammate:
            stats_by_teammate[uid] = {
                'teammate_id': uid,
                'teammate_name': gp.user.get_display_name(),
                'teammate_username': gp.user.username,
                'game_count': 0,
                'wins': 0,
                'losses': 0,
                'total_score': 0,
                'last_played': None,
            }

        my_gp = my_scores_by_game.get(gp.game_id)
        if my_gp:
            s = stats_by_teammate[uid]
            s['game_count'] += 1
            s['total_score'] += my_gp.score
            if my_gp.is_winner:
                s['wins'] += 1
            else:
                s['losses'] += 1
            if s['last_played'] is None or gp.game.game_time > s['last_played']:
                s['last_played'] = gp.game.game_time

    result = []
    for s in stats_by_teammate.values():
        game_count = s['game_count']
        avg_score = round(s['total_score'] / game_count, 2) if game_count > 0 else 0
        win_rate = round(s['wins'] / game_count * 100, 2) if game_count > 0 else 0

        result.append({
            'teammate_id': s['teammate_id'],
            'teammate_name': s['teammate_name'],
            'teammate_username': s['teammate_username'],
            'game_count': game_count,
            'avg_score': avg_score,
            'win_rate': win_rate,
            'wins': s['wins'],
            'losses': s['losses'],
            'total_score': s['total_score'],
            'last_played': s['last_played'].isoformat() if s['last_played'] else None,
        })

    sort_map = {
        'game_count': 'game_count',
        'avg_score': 'avg_score',
        'win_rate': 'win_rate',
        'total_score': 'total_score',
        'last_played': 'last_played',
    }
    sort_field = sort_map.get(sort_by, 'game_count')

    reverse = order == 'desc'
    result.sort(key=lambda x: (x[sort_field] is None, x[sort_field]), reverse=reverse)

    return result[:limit]


RESTORE_TYPE_RECOVERABLE = 'recoverable'
RESTORE_TYPE_DUPLICATE = 'duplicate'
RESTORE_TYPE_MISSING_USER = 'missing_user'
RESTORE_TYPE_FORMAT_ERROR = 'format_error'

RESTORE_TYPE_LABELS = {
    RESTORE_TYPE_RECOVERABLE: '可恢复记录',
    RESTORE_TYPE_DUPLICATE: '重复战绩',
    RESTORE_TYPE_MISSING_USER: '缺失用户',
    RESTORE_TYPE_FORMAT_ERROR: '格式冲突',
}

RESTORE_TYPE_COLORS = {
    RESTORE_TYPE_RECOVERABLE: 'success',
    RESTORE_TYPE_DUPLICATE: 'info',
    RESTORE_TYPE_MISSING_USER: 'warning',
    RESTORE_TYPE_FORMAT_ERROR: 'danger',
}


class RestoreItem:
    def __init__(self, index, game_data=None, message='', details=None):
        self.index = index
        self.game_data = game_data or {}
        self.message = message
        self.details = details or {}

    def to_dict(self):
        return {
            'index': self.index,
            'game_data': self.game_data,
            'message': self.message,
            'details': self.details,
        }


class RestoreResult:
    def __init__(self):
        self.recoverable = []
        self.duplicates = []
        self.missing_users = []
        self.format_errors = []
        self.total_count = 0

    def add_recoverable(self, item):
        self.recoverable.append(item)
        self.total_count += 1

    def add_duplicate(self, item):
        self.duplicates.append(item)
        self.total_count += 1

    def add_missing_user(self, item):
        self.missing_users.append(item)
        self.total_count += 1

    def add_format_error(self, item):
        self.format_errors.append(item)
        self.total_count += 1

    @property
    def recoverable_count(self):
        return len(self.recoverable)

    @property
    def duplicate_count(self):
        return len(self.duplicates)

    @property
    def missing_user_count(self):
        return len(self.missing_users)

    @property
    def format_error_count(self):
        return len(self.format_errors)

    def get_category_summary(self):
        categories = []
        for type_key, label in RESTORE_TYPE_LABELS.items():
            if type_key == RESTORE_TYPE_RECOVERABLE:
                items = self.recoverable
            elif type_key == RESTORE_TYPE_DUPLICATE:
                items = self.duplicates
            elif type_key == RESTORE_TYPE_MISSING_USER:
                items = self.missing_users
            else:
                items = self.format_errors

            if items:
                categories.append({
                    'type': type_key,
                    'label': label,
                    'count': len(items),
                    'color': RESTORE_TYPE_COLORS.get(type_key, 'secondary'),
                    'samples': [item.to_dict() for item in items[:5]],
                })
        return categories

    def get_summary_dict(self):
        return {
            'total': self.total_count,
            'recoverable_count': self.recoverable_count,
            'duplicate_count': self.duplicate_count,
            'missing_user_count': self.missing_user_count,
            'format_error_count': self.format_error_count,
            'categories': self.get_category_summary(),
        }

    def get_all_errors(self):
        errors = []
        for item in self.duplicates:
            errors.append(item.to_dict())
        for item in self.missing_users:
            errors.append(item.to_dict())
        for item in self.format_errors:
            errors.append(item.to_dict())
        return errors


def _get_game_signature(game_time, players_data):
    """生成游戏的唯一签名，用于检测重复战绩

    使用 游戏时间 + 排序后的玩家用户名集合 + 排序后的得分集合 作为签名
    """
    from datetime import datetime
    if isinstance(game_time, datetime):
        time_str = game_time.strftime('%Y-%m-%d %H:%M')
    else:
        time_str = str(game_time)

    usernames = sorted(p.get('username', '') for p in players_data)
    scores = sorted(str(p.get('score', 0)) for p in players_data)
    return f"{time_str}|{','.join(usernames)}|{','.join(scores)}"


def analyze_backup_file(file_obj):
    """分析备份文件，将记录分类为可恢复、重复、缺失用户和格式冲突

    Args:
        file_obj: 上传的备份文件对象

    Returns:
        RestoreResult: 结构化的恢复分析结果
    """
    import json
    from datetime import datetime
    from django.utils import timezone
    from apps.accounts.models import User
    from .models import Game, GamePlayer

    result = RestoreResult()

    try:
        data = json.loads(file_obj.read().decode('utf-8'))
        games_data = data.get('games', [])
        if not isinstance(games_data, list):
            raise ValueError('games 字段必须是数组')
    except Exception as e:
        result.add_format_error(RestoreItem(
            index=0,
            message=f'备份文件解析失败：{e}',
            details={'exception': str(e)},
        ))
        return result

    existing_signatures = set()
    for game in Game.objects.filter(status='completed').prefetch_related('players__user'):
        players_data = [
            {'username': gp.user.username, 'score': gp.score}
            for gp in game.players.all()
        ]
        sig = _get_game_signature(game.game_time, players_data)
        existing_signatures.add(sig)

    for i, gd in enumerate(games_data, 1):
        try:
            if not isinstance(gd, dict):
                result.add_format_error(RestoreItem(
                    index=i,
                    message=f'第{i}条：数据格式错误，不是对象',
                    details={'raw_type': type(gd).__name__},
                ))
                continue

            game_time_str = gd.get('game_time', '')
            players_raw = gd.get('players', [])

            if not game_time_str:
                result.add_format_error(RestoreItem(
                    index=i,
                    message=f'第{i}条：缺少游戏时间',
                    details={'game_data': gd},
                ))
                continue

            if not players_raw or not isinstance(players_raw, list):
                result.add_format_error(RestoreItem(
                    index=i,
                    message=f'第{i}条：玩家数据格式错误',
                    details={'players_type': type(players_raw).__name__},
                ))
                continue

            try:
                gt = datetime.fromisoformat(game_time_str)
                if gt.tzinfo is None:
                    gt = timezone.make_aware(gt)
            except (ValueError, TypeError):
                result.add_format_error(RestoreItem(
                    index=i,
                    message=f'第{i}条：时间格式错误 "{game_time_str}"',
                    details={'game_time_str': game_time_str},
                ))
                continue

            players_data = []
            missing_usernames = []
            has_format_error = False
            total_score = 0

            for pd in players_raw:
                if not isinstance(pd, dict):
                    has_format_error = True
                    continue

                username = pd.get('username', '')
                try:
                    score = int(pd.get('score', 0))
                except (ValueError, TypeError):
                    has_format_error = True
                    continue

                if not username:
                    has_format_error = True
                    continue

                try:
                    user = User.objects.get(username=username)
                    players_data.append({
                        'user': user,
                        'username': username,
                        'score': score,
                        'is_winner': pd.get('is_winner', score > 0),
                    })
                    total_score += score
                except User.DoesNotExist:
                    missing_usernames.append(username)

            if has_format_error:
                result.add_format_error(RestoreItem(
                    index=i,
                    message=f'第{i}条：玩家数据格式错误',
                    details={'players_raw': players_raw},
                ))
                continue

            if missing_usernames:
                result.add_missing_user(RestoreItem(
                    index=i,
                    game_data={
                        'game_time': gt.isoformat(),
                        'location': gd.get('location', ''),
                        'game_type': gd.get('game_type', ''),
                        'players': [{'username': p.get('username', ''), 'score': p.get('score', 0)} for p in players_raw],
                    },
                    message=f'第{i}条：用户 {", ".join(missing_usernames)} 不存在',
                    details={'missing_usernames': missing_usernames},
                ))
                continue

            if len(players_data) < 2:
                result.add_format_error(RestoreItem(
                    index=i,
                    message=f'第{i}条：玩家数量不足（至少2人）',
                    details={'player_count': len(players_data)},
                ))
                continue

            if total_score != 0:
                result.add_format_error(RestoreItem(
                    index=i,
                    message=f'第{i}条：得分总和不为0（当前为{total_score}）',
                    details={'total_score': total_score},
                ))
                continue

            sig_players = [{'username': p['username'], 'score': p['score']} for p in players_data]
            signature = _get_game_signature(gt, sig_players)

            if signature in existing_signatures:
                result.add_duplicate(RestoreItem(
                    index=i,
                    game_data={
                        'game_time': gt.isoformat(),
                        'location': gd.get('location', ''),
                        'game_type': gd.get('game_type', ''),
                        'base_score': gd.get('base_score', 1),
                        'players': [
                            {'username': p['username'], 'score': p['score']}
                            for p in players_data
                        ],
                    },
                    message=f'第{i}条：与现有战绩重复',
                    details={'signature': signature},
                ))
                continue

            creator = None
            creator_username = gd.get('creator', '')
            if creator_username:
                creator = User.objects.filter(username=creator_username).first()

            result.add_recoverable(RestoreItem(
                index=i,
                game_data={
                    'game_time': gt.isoformat(),
                    'location': gd.get('location', ''),
                    'game_type': gd.get('game_type', 'mahjong_16'),
                    'base_score': gd.get('base_score', 1),
                    'notes': gd.get('notes', ''),
                    'is_supplemental': gd.get('is_supplemental', True),
                    'creator_username': creator_username,
                    'creator_exists': creator is not None,
                    'players': [
                        {
                            'username': p['username'],
                            'display_name': p['user'].get_display_name(),
                            'score': p['score'],
                            'is_winner': p['is_winner'],
                        }
                        for p in players_data
                    ],
                },
                message=f'第{i}条：可恢复',
                details={'players_count': len(players_data)},
            ))

        except Exception as e:
            result.add_format_error(RestoreItem(
                index=i,
                message=f'第{i}条：解析错误 - {e}',
                details={'exception': str(e)},
            ))

    return result
